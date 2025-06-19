
import os
import warnings
import json
import re
from PyPDF2 import PdfReader, PdfWriter
from dotenv import load_dotenv
from langchain_community.document_loaders import UnstructuredFileLoader
import google.generativeai as genai
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from openpyxl import load_workbook
import pandas as pd
import shutil
import threading
from datetime import datetime
from dateutil.relativedelta import relativedelta

warnings.filterwarnings("ignore")
load_dotenv()

base_dir = os.path.dirname(os.path.abspath(__file__))
# ## Configure Gemini API
genai.configure(api_key=os.getenv("GOOGLE_API_KEY_V2"))
model = genai.GenerativeModel("gemini-1.5-flash")


# ## Folders
input_folder = os.path.join(base_dir, '..', 'temp_combine')
output_folder = os.path.join(base_dir, '..', 'temp_out_combine')
os.makedirs(output_folder, exist_ok=True)

# Thread lock for Excel writing
excel_lock = threading.Lock()

# ## Extract First Page from PDF
def extract_first_page(input_path, output_path):
    # try:
    #     reader = PdfReader(input_path)
    #     writer = PdfWriter()
    #     if reader.pages:
    #         writer.add_page(reader.pages[0])
    #         with open(output_path, 'wb') as f_out:
    #             writer.write(f_out)
    #     return output_path
    # except Exception as e:
    #     return f"Error processing {input_path}: {str(e)}"
    try:
        reader = PdfReader(input_path)
        writer = PdfWriter()
        num_pages_to_extract = min(2, len(reader.pages))  # Avoid IndexError if file has less than 2 pages

        for i in range(num_pages_to_extract):
            writer.add_page(reader.pages[i])

        with open(output_path, 'wb') as f_out:
            writer.write(f_out)

        return output_path
    except Exception as e:
        return f"Error processing {input_path}: {str(e)}"

# ## Extract text from a single PDF using OCR
def extract_text_from_pdf(input_path):
    loader = UnstructuredFileLoader(input_path, mode="elements", strategy="ocr_only")
    docs = loader.load()
    return '\n'.join([doc.page_content for doc in docs])

# ## Read Prompt
def get_prompt():
    prompt_path = os.path.join(base_dir, 'common_prompt.txt')
    with open(prompt_path, 'r') as f:
        return f.read()

# ## Extract KPIs using Gemini
def extract_kpis_with_gemini(prompt, invoice_text):
    full_prompt = f"{prompt}\n\n📄 Here is the invoice text:\n\n{invoice_text}"
    try:
        response = model.generate_content(full_prompt,
                                          generation_config={
                                                "temperature": 0.0
                                            })
        match = re.search(r"\{.*\}", response.text, re.DOTALL)
        return json.loads(match.group()) if match else {"error": "No JSON found in response."}
    except Exception as e:
        return {"error": f"Gemini error: {str(e)}"}

# ## Pipeline for a single PDF file
def process_pdf_file(filename, prompt):
    try:
        input_path = os.path.join(input_folder, filename)
        output_path = os.path.join(output_folder, filename)

        # Step 1: Extract first page
        extract_first_page(input_path, output_path)

        # Step 2: OCR text extraction
        invoice_text = extract_text_from_pdf(output_path)
        if not invoice_text.strip():
            return filename, {"error": "Empty or invalid text extracted."}

        # Step 3: Extract KPIs using Gemini
        result = extract_kpis_with_gemini(prompt, invoice_text)
        return filename, result

    except Exception as e:
        return filename, {"error": str(e)}

# ## Run Pipeline - OPTION 2 (Recommended: Collect all data first, then write)

def get_date_cols(billing_dates_str):
    print("Before processing : Billing Dates String:", billing_dates_str)

#     converted_dates = [
#     f"{date_obj.month}/{date_obj.strftime('%d/%y')}"
#     for date_obj in [datetime.strptime(d, '%m/%d/%y') for d in billing_dates_str]   
# ]
    # print("After Processing --------Converted Billing Dates String:", converted_dates)
    
    # Convert to datetime objects
    billing_dates = sorted([datetime.strptime(date, '%m/%d/%y') for date in billing_dates_str])
    print("After osrting ----Converted Billing Dates:", billing_dates)
    # Initialize result list
    data = []
    
    result_dict = {}
    # Loop through billing dates to generate 'From' and 'To'
    for i in range(len(billing_dates)):
        to_date = billing_dates[i]
        if i == 0:
            from_date = to_date - relativedelta(months=1)
        else:
            from_date = billing_dates[i - 1]
        
        print("date :", to_date, "from:", from_date)
        # Calculate number of days (inclusive of both ends)
        no_of_days = abs((to_date - from_date).days ) # add 1 if you want to include both start and end dates

        # Format dates manually without leading zeros
        billing_str = f"{to_date.month}/{to_date.day}/{to_date.strftime('%y')}"
        from_str = f"{from_date.month}/{from_date.day}/{from_date.strftime('%y')}"
        to_str = billing_str

       

        # Store using the original unsorted full format string (to match input)
        original_key = to_date.strftime('%m/%d/%y')
        print('original_key-------------->>>', original_key)

        result_dict[original_key] = {
            'From': from_str,
            'To': to_str
        }

    # # Convert to DataFrame
    # df = pd.DataFrame(data)
    # return df
    return result_dict

# Define regex pattern for '8 February 2025' format
pattern = r"^\d{1,2} (January|February|March|April|May|June|July|August|September|October|November|December) \d{4}$"
# Example: Apply logic only to matched values (e.g., convert to ISO format)
def convert_date(value):
    if re.match(pattern, str(value)):
        return pd.to_datetime(value, errors='coerce').strftime("%m/%d/%y")
    else:
        return value  # or return pd.NaT / None


# Clean numeric columns (remove commas, currency symbols, convert to float)
def clean_numeric(val):
    if isinstance(val, str):
        val = val.replace(",", "").replace("£", "").replace("$", "").strip()
        try:
            return float(val)
        except ValueError:
            return None
    return val

# Alternative approach: Collect all data first, then write once
def write_all_kpis_to_excel(template_path, output_path, all_kpi_data):
    """
    Alternative function to write all KPI data at once instead of individually
    """
    # Define regex pattern for '8 February 2025' format
    pattern = r"^\d{1,2} (January|February|March|April|May|June|July|August|September|October|November|December) \d{4}$"

    if not all_kpi_data:
        print("No valid KPI data to write")
        return
    
    # Convert all data to DataFrame
    df = pd.DataFrame(all_kpi_data)

    df['Billing Date'] = df['Billing Date'].apply(convert_date)
    df["Billing Date"] = pd.to_datetime(df["Billing Date"], errors='coerce')

    df["Month"] = df["Billing Date"].dt.strftime("%b-%y")
    df["Billing Date"] = pd.to_datetime(df["Billing Date"], errors='coerce').dt.strftime("%m/%d/%y")

    df['From'] = pd.to_datetime(df['From'], dayfirst=True).dt.strftime('%m/%d/%y')
    df['To'] = pd.to_datetime(df['To'], dayfirst=True).dt.strftime('%m/%d/%y')


    
    # for col in ["To"]:
    #     if col in df.columns:
    #         df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime("%m/%d/%y")

    billing_dates = df.loc[df['From'].isna() & df['To'].isna(), 'Billing Date'].tolist()
    print('>>>>>>>>>>>>>>', billing_dates)
    date_info_dict = {}
    if billing_dates:
        date_info_dict = get_date_cols(billing_dates)

    if date_info_dict:
        matched_dates = df['Billing Date'].isin(date_info_dict.keys())

        df.loc[matched_dates, 'From'] = df.loc[matched_dates, 'Billing Date'].map(
            lambda x: date_info_dict[x].get('From')
        )
        df.loc[matched_dates, 'To'] = df.loc[matched_dates, 'Billing Date'].map(
            lambda x: date_info_dict[x].get('To')
        )

  
    numeric_cols = [
        "Day kWh", "Night kWh", "DUoS Capacity Charge", "Excess Capacity Charge", "VAT", 
        "Total Invoice value"
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)

    # # Derived columns
    # if all(col in df.columns for col in ["Billing Date", "From", "To"]):
    #     # Convert back to datetime for calculations, then format for display
    #     # billing_dt = pd.to_datetime(df["Billing Date"], format="%d/%m/%y", errors='coerce')
    #     df["Month"] = df["Billing Date"].dt.strftime("%b-%y")
    #     df["Billing Date"] = pd.to_datetime(df["Billing Date"], errors='coerce').dt.strftime("%m/%d/%y")

    if all(col in df.columns for col in ["Billing Date", "From", "To"]):
    
        df['No of Days'] = (
                        pd.to_datetime(df['To'], format='%m/%d/%y') - 
                        pd.to_datetime(df['From'], format='%m/%d/%y')
                    ).dt.days

    print('>>>>>>>>>>>>>>>Invoice value', df['Total Invoice value'].tolist())
    # else:
    #     df["Month"] = None
    #     df["No of Days"] = None

    # if all(col in df.columns for col in ["Day kWh", "Night kWh", "No of Days"]):
    #     df["kWh per day"] = (df["Day kWh"] + df["Night kWh"]) / df["No of Days"].replace(0, 1)
    #     df["kWh per day"] = df["kWh per day"].round(1)
        
    # else:
    #     df["kWh per day"] = None

    if all(col in df.columns for col in ["Total Invoice value", "VAT"]):
        df["Total $ amount (Without VAT)"] = df["Total Invoice value"] - df["VAT"]


       
    # else:
    #     df["Total $ amount (Without VAT)"] = None

    total_kwh = None
    if all(col in df.columns for col in ["Day kWh", "Night kWh"]):
        total_kwh = df["Day kWh"] + df["Night kWh"]
        df["Total Kwh"] = total_kwh
        
    # Check if all columns exist
    has_only_peak = all(col in df.columns for col in ["Usage", "On-Peak"])
    has_total_kwh = "Total Kwh" in df.columns

    if has_total_kwh and has_only_peak:
        
        df.loc[df['Total Kwh'].isna(), 'Total Kwh'] = df.loc[df['Total Kwh'].isna(), 'Usage']
        # df.loc[df['Total Kwh'].isna(), 'Total Kwh'] = (
        #                                             df['Only kWh'] + df["On-Peak"]
        #                                         )
        # df["Total Kwh"] = df["Total Kwh"].fillna(df["Only kWh"] + df["On-Peak"])
    elif has_only_peak:
        df["Total Kwh"] = df["Usage"] + df["On-Peak"]

    total_kwh = df["Total Kwh"].sum()
    
    if  "Total Invoice value" in df.columns:
        total_kwh = df["Total Kwh"].sum()
        # safe_total_kwh = total_kwh if total_kwh != 0 else 1
        # df["Blended rate $/kWh (With VAT)"] = df["Total Invoice value"] / df["Total Kwh"]
        df.loc[df['Total Invoice value'].notna(), 'Blended rate $/kWh (With VAT)'] = df.loc[df['Total Invoice value'].notna(), 'Total Invoice value'] /df.loc[df['Total Invoice value'].notna(), 'Total Kwh']
        df["Blended rate $/kWh (With VAT)"] = df["Blended rate $/kWh (With VAT)"].round(2)
    else:
        df["Blended rate $/kWh (With VAT)"] = None

    if total_kwh is not None and "Total $ amount (Without VAT)" in df.columns:
        # df["Blended rate $/kWh (Without VAT)"] = df["Total $ amount (Without VAT)"] / df["Total Kwh"]
        df.loc[df['Total $ amount (Without VAT)'].notna(), 'Blended rate $/kWh (Without VAT)'] = df.loc[df['Total $ amount (Without VAT)'].notna(), 'Total $ amount (Without VAT)'] /df.loc[df['Total $ amount (Without VAT)'].notna(), 'Total Kwh']

        df["Blended rate $/kWh (Without VAT)"] = df["Blended rate $/kWh (Without VAT)"].round(2)    
    else:
        df["Blended rate $/kWh (Without VAT)"] = None

    if "Current Electric Charges" in df.columns:
        df['Total $ amount'] = df['Current Electric Charges'] + df['Generation/Retail $ amount'].fillna(0)

    if "City of Cullman Tax $" in df.columns:
            df["City of Cullman Tax $ (2)"] = (df['City of Cullman Tax $']) / (df['Total $ amount'] - df['City of Cullman Tax $']) * 100
            df['City of Cullman Tax $ (2)'] = df['City of Cullman Tax $ (2)'].round(2)
    
    if "Alabama State Taxes $" in df.columns:
        df["Alabama State Taxes $ (2)"] = (df['Alabama State Taxes $']) / (df['Total $ amount'] - df['Alabama State Taxes $']) * 100
        df['Alabama State Taxes $ (2)'] = df['Alabama State Taxes $ (2)'].round(2)

    if "Total Kwh" in df.columns:
        df["kWh per day"] = df["Total Kwh"] / df["No of Days"]
        df["kWh per day"] = df["kWh per day"].round(2)

    if "Total Kwh" in df.columns:
        df["Blended rate $/kWh"] = df["Total $ amount"]/df["Total Kwh"]
        df["Blended rate $/kWh"] = df["Blended rate $/kWh"].round(3)

    if 'Total Kwh' in df.columns:
            total_kwh = df['Total Kwh'].fillna(0).sum()
            df['kWh %'] = (df['Total Kwh'] / total_kwh * 100).round(0)
            # df['kWh %'] = df['kWh %'].round(0)

    if "VAT" in df.columns:
        df['VAT'] = df['VAT'].round(2)


    # Column mapping
    column_map = {
        "Billing Date": 2, "Month": 3, "From": 4, "To": 5, "No of Days": 6,
        "Day kWh": 7, "Night kWh": 8, "Usage": 13, "Total Kwh": 14, "kWh per day": 15,
        "kWh %": 16, "DUoS Capacity Charge": 17, "Excess Capacity Charge": 18,
        "VAT": 23, "City of Cullman Tax $":27, "Alabama State Taxes $": 28, "City of Cullman Tax $ (2)":30, "Alabama State Taxes $ (2)":29, "Total Invoice value": 33, "Total $ amount (Without VAT)": 34,
        "Current Electric Charges": 35,
        "Total $ amount": 36,
        "Blended rate $/kWh (With VAT)": 38, "Blended rate $/kWh (Without VAT)": 39, "Blended rate $/kWh": 40
    }
    
    # Copy template and write all data
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # Write all records starting from row 3
    start_row = 3
    for idx, (_, record) in enumerate(df.iterrows()):
        current_row = start_row + idx
        for kpi, col in column_map.items():
            value = record.get(kpi, "")
            ws.cell(row=current_row, column=col).value = value

    wb.save(output_path)
    print(f"✅ All {len(df)} records written to {output_path}")


def novolex_rwc_run_pipeline_batch_write():
    prompt = get_prompt()
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    pdf_files.sort(key=lambda x: len(x)) 
    template_path = os.path.join(base_dir, '..', 'RedaptiveCombinedFromate.xlsx')
    output_excel_path = os.path.join(base_dir,  '..', 'combined_output.xlsx')

    results = {}
    valid_kpi_data = []
    with ThreadPoolExecutor(max_workers=4) as executor:
        future_to_file = {executor.submit(process_pdf_file, f, prompt): f for f in pdf_files}
        for future in as_completed(future_to_file):
            file = future_to_file[future]
            try:
                fname, res = future.result()
                results[fname] = res
                print(f"📄 Processed ----------->>>>>{fname}: {res}")
                
                # Collect valid KPI data
                if isinstance(res, dict) and not res.get("error"):
                    res['filename'] = fname  # Add filename for reference
                    valid_kpi_data.append(res)
                else:
                    print(f"⚠️ Skipping {fname}: {res.get('error')}")

            except Exception as e:
                results[file] = {"error": f"Unhandled exception: {str(e)}"}
    # Write all valid data to Excel at once
    if valid_kpi_data:
        write_all_kpis_to_excel(template_path, output_excel_path, valid_kpi_data)
    
    return results


# ## Execute
if __name__ == "__main__":
    start_time = time.time()
    
    # Use OPTION 1 (fixed original) or OPTION 2 (batch write - recommended)
    # all_results = run_pipeline()  # Option 1
    all_results = novolex_rwc_run_pipeline_batch_write()  # Option 2 - Recommended
    
    end_time = time.time()
    total_time = end_time - start_time
    for file, result in all_results.items():
        print(f"\n🧾 File: {file}\n📊 KPI Result: {result}")

    print(f"\n⏱️ Total processing time: {total_time:.2f} seconds")


